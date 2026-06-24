import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, cast
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.models import (
    CheckSession,
    CheckItem,
    MatchResult,
    SourceItem,
    SourceDocument,
)


def format_value_unit(val: float, unit: str) -> str:
    """Formats float values cleanly by removing trailing zeros and adding unit."""
    if val == int(val):
        formatted_val = f"{int(val):,}"
    else:
        val_str = f"{val:f}"
        if "." in val_str:
            val_str = val_str.rstrip("0").rstrip(".")
        if "." in val_str:
            parts = val_str.split(".")
            formatted_val = f"{int(parts[0]):,}.{parts[1]}"
        else:
            formatted_val = f"{int(val_str):,}"
    return f"{formatted_val} {unit}"


def format_item_value(item: Any) -> str:
    """value_type に応じたアイテム値の表示文字列を返す."""
    vtype = getattr(item, "value_type", "numeric") or "numeric"
    if vtype == "numeric":
        val = float(item.value) if item.value is not None else 0.0
        unit = str(item.unit) if item.unit else ""
        return format_value_unit(val, unit)
    elif vtype == "name":
        return str(item.text_value) if item.text_value else ""
    elif vtype == "formula":
        return str(item.formula_value) if item.formula_value else ""
    return ""


def generate_excel_report(db: Session, session_id: str) -> io.BytesIO:
    """Generates a styled Excel verification report for the given CheckSession."""
    session = db.query(CheckSession).filter(CheckSession.id == session_id).first()
    if not session:
        raise ValueError(f"Session with ID {session_id} not found")

    check_items = db.query(CheckItem).filter(CheckItem.session_id == session_id).all()
    check_item_ids = [item.id for item in check_items]
    match_results = (
        db.query(MatchResult)
        .filter(MatchResult.check_item_id.in_(check_item_ids))
        .all()
    )

    # Dynamic annotation ID mapping (matches PDF annotator logic)
    matched_ids: dict[str, str] = {}
    matched_idx = 1
    for r in match_results:
        if r.status == "approved":
            matched_ids[cast(str, r.check_item_id)] = f"c{matched_idx:02d}"
            matched_idx += 1

    # Map database records to structured rows
    results_lookup = {cast(str, r.check_item_id): r for r in match_results}
    results_data: list[dict] = []

    for item in check_items:
        r = results_lookup.get(cast(str, item.id))
        annotation_id = "-"
        status_str = "Pending"
        ai_reason = "No matching guideline found in reference documents."
        source_filename = "-"
        source_val_unit = "-"
        source_page_str = "-"

        if r:
            if r.status == "approved":
                annotation_id = matched_ids.get(cast(str, item.id), "-")
                status_str = "Approved"
            elif r.status == "rejected":
                status_str = "Mismatch"
            else:
                status_str = "Pending"

            ai_reason = r.ai_reasoning or ""

            if r.source_item_id:
                s_item = (
                    db.query(SourceItem)
                    .filter(SourceItem.id == r.source_item_id)
                    .first()
                )
                if s_item:
                    s_doc = (
                        db.query(SourceDocument)
                        .filter(SourceDocument.id == s_item.document_id)
                        .first()
                    )
                    if s_doc:
                        source_filename = cast(str, s_doc.filename)
                    source_val_unit = format_item_value(s_item)
                    source_page_str = str(s_item.page)

        results_data.append(
            {
                "annotation_id": annotation_id,
                "status": status_str,
                "item_label": item.label,
                "calc_val_unit": format_item_value(item),
                "calc_page": str(item.page),
                "source_filename": source_filename,
                "source_val_unit": source_val_unit,
                "source_page": source_page_str,
                "ai_reasoning": ai_reason,
            }
        )

    # Create OpenPyXL workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MarkVal Report"

    # Configure print settings for A4 Landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Style templates
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    font_meta = Font(name="Calibri", size=10, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)

    fill_header = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    fill_approved = PatternFill(
        start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"
    )
    fill_mismatch = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    fill_pending = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Write brand and metadata headers
    ws["A1"] = "MarkVal Co-pilot Verification Report"
    ws["A1"].font = font_title

    filename_a = Path(cast(str, session.file_a_path)).name
    now_str = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S")

    ws["A2"] = f"Session ID: {session_id}"
    ws["A2"].font = font_meta
    ws["A3"] = f"Target File A: {filename_a}"
    ws["A3"].font = font_meta
    ws["A4"] = f"Generated At: {now_str} (Local Time)"
    ws["A4"].font = font_meta

    # Write main data headers (optimized for A4 landscape print)
    headers = [
        "ID",
        "AI判定",
        "項目名",
        "計算値",
        "頁(A)",
        "出典ファイル",
        "出典値",
        "頁(B)",
        "AI理由",
        "目視確認",
        "確認者",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = border_thin

    ws.row_dimensions[6].height = 28

    # Write data rows
    current_row = 7
    for item in results_data:
        row_data = [
            item["annotation_id"],
            item["status"],
            item["item_label"],
            item["calc_val_unit"],
            item["calc_page"],
            item["source_filename"],
            item["source_val_unit"],
            item["source_page"],
            item["ai_reasoning"],
            "",  # Human verification (empty)
            "",  # Checker name (empty)
        ]

        status = item["status"]
        if status == "Approved":
            row_fill = fill_approved
        elif status == "Mismatch":
            row_fill = fill_mismatch
        else:
            row_fill = fill_pending

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin

            # Apply background color only to AI-generated columns (1 to 9).
            # Columns 10 & 11 (human checks) remain white.
            if col_idx <= 9:
                cell.fill = row_fill

            # Align columns
            if col_idx in [1, 2, 5, 8, 10, 11]:
                cell.alignment = align_center
            elif col_idx in [4, 7]:
                cell.alignment = align_right
            else:
                cell.alignment = align_left_wrap

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Adjust column widths dynamically with limits
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)

        # Skip metadata rows when calculating widths
        max_len = 0
        for cell in col[5:]:
            if cell.value:
                # Japanese characters take double width roughly
                val_str = str(cell.value)
                length = sum(2 if ord(char) > 255 else 1 for char in val_str)
                max_len = max(max_len, length)

        # Precise constraints for print limits
        if col_letter == "I":  # AI理由
            ws.column_dimensions[col_letter].width = 45
        elif col_letter == "F":  # 出典ファイル
            ws.column_dimensions[col_letter].width = 25
        elif col_letter == "C":  # 項目名
            ws.column_dimensions[col_letter].width = 25
        elif col_letter == "A":  # ID
            ws.column_dimensions[col_letter].width = 8
        elif col_letter == "B":  # AI判定
            ws.column_dimensions[col_letter].width = 12
        elif col_letter == "E":  # 頁(A)
            ws.column_dimensions[col_letter].width = 8
        elif col_letter == "H":  # 頁(B)
            ws.column_dimensions[col_letter].width = 8
        elif col_letter in ["J", "K"]:  # 目視確認, 確認者
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save workbook to memory stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
